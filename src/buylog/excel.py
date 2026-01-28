#!/usr/bin/env python3
"""
Excel import/export functionality for buylog.

Supports reading and writing Excel files for:
- Vendors, Products, Brands
- Quotes, Purchase Orders
- Specifications with features
- Forex rates, Price alerts, Watchlists
"""

import datetime
from pathlib import Path
from typing import Any, Optional, Union

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from sqlalchemy.orm import Session

from .models import PO_STATUSES, SPEC_DATA_TYPES
from .services import (
    BrandService,
    ProductService,
    VendorService,
    QuoteService,
    SpecificationService,
    PurchaseOrderService,
    DuplicateError,
    NotFoundError,
)


# Styling constants
HEADER_FONT = Font(bold=True, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


def _style_header(ws, num_cols: int) -> None:
    """Apply header styling to first row."""
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER


def _auto_width(ws) -> None:
    """Auto-adjust column widths based on content."""
    for column_cells in ws.columns:
        max_length = 0
        column = column_cells[0].column_letter
        for cell in column_cells:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except Exception:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column].width = adjusted_width


def _add_data_validation(
    ws, col: int, values: list[str], start_row: int = 2, end_row: int = 1000
) -> None:
    """Add dropdown data validation to a column."""
    dv = DataValidation(
        type="list",
        formula1=f'"{",".join(values)}"',
        allow_blank=True,
    )
    dv.error = "Please select from the list"
    dv.errorTitle = "Invalid entry"
    col_letter = get_column_letter(col)
    dv.add(f"{col_letter}{start_row}:{col_letter}{end_row}")
    ws.add_data_validation(dv)


# =============================================================================
# Export Functions
# =============================================================================


def export_brands(session: Session, filepath: Union[str, Path]) -> int:
    """Export all brands to Excel. Returns number of rows exported."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Brands"

    headers = ["ID", "Name", "Product Count"]
    ws.append(headers)
    _style_header(ws, len(headers))

    brands = BrandService.get_all(session)
    for brand in brands:
        ws.append([brand.id, brand.name, len(brand.products)])

    _auto_width(ws)
    wb.save(filepath)
    return len(brands)


def export_products(session: Session, filepath: Union[str, Path]) -> int:
    """Export all products to Excel. Returns number of rows exported."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Products"

    headers = ["ID", "Name", "Brand", "Category", "Specification", "Quote Count"]
    ws.append(headers)
    _style_header(ws, len(headers))

    products = ProductService.get_all(session)
    for product in products:
        ws.append(
            [
                product.id,
                product.name,
                product.brand.name if product.brand else "",
                product.category or "",
                product.specification.name if product.specification else "",
                len(product.quotes),
            ]
        )

    _auto_width(ws)
    wb.save(filepath)
    return len(products)


def export_vendors(session: Session, filepath: Union[str, Path]) -> int:
    """Export all vendors to Excel. Returns number of rows exported."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Vendors"

    headers = [
        "ID",
        "Name",
        "Currency",
        "Discount Code",
        "Discount %",
        "URL",
        "Contact Person",
        "Email",
        "Phone",
        "Website",
        "Address Line 1",
        "Address Line 2",
        "City",
        "State",
        "Postal Code",
        "Country",
        "Tax ID",
        "Payment Terms",
    ]
    ws.append(headers)
    _style_header(ws, len(headers))

    vendors = VendorService.get_all(session)
    for v in vendors:
        ws.append(
            [
                v.id,
                v.name,
                v.currency,
                v.discount_code,
                v.discount,
                v.url,
                v.contact_person,
                v.email,
                v.phone,
                v.website,
                v.address_line1,
                v.address_line2,
                v.city,
                v.state,
                v.postal_code,
                v.country,
                v.tax_id,
                v.payment_terms,
            ]
        )

    _auto_width(ws)
    wb.save(filepath)
    return len(vendors)


def export_quotes(session: Session, filepath: Union[str, Path]) -> int:
    """Export all quotes to Excel. Returns number of rows exported."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Quotes"

    headers = [
        "ID",
        "Vendor",
        "Product",
        "Brand",
        "Price",
        "Currency",
        "Shipping Cost",
        "Total Cost",
        "Status",
        "Created At",
    ]
    ws.append(headers)
    _style_header(ws, len(headers))

    quotes = QuoteService.get_all(session)
    for q in quotes:
        ws.append(
            [
                q.id,
                q.vendor.name,
                q.product.name,
                q.product.brand.name if q.product.brand else "",
                q.value,
                q.currency,
                q.shipping_cost,
                q.total_cost,
                q.status,
                q.created_at.isoformat() if q.created_at else "",
            ]
        )

    _auto_width(ws)
    wb.save(filepath)
    return len(quotes)


def export_specifications(session: Session, filepath: Union[str, Path]) -> int:
    """Export all specifications to Excel. Returns number of rows exported."""
    wb = Workbook()

    # Sheet 1: Specifications
    ws_specs = wb.active
    ws_specs.title = "Specifications"
    spec_headers = ["ID", "Name", "Description", "Feature Count", "Product Count"]
    ws_specs.append(spec_headers)
    _style_header(ws_specs, len(spec_headers))

    specs = SpecificationService.get_all(session)
    for spec in specs:
        ws_specs.append(
            [
                spec.id,
                spec.name,
                spec.description or "",
                len(spec.features),
                len(spec.products),
            ]
        )
    _auto_width(ws_specs)

    # Sheet 2: Features
    ws_features = wb.create_sheet("Features")
    feature_headers = [
        "Specification",
        "Feature Name",
        "Data Type",
        "Unit",
        "Required",
        "Min Value",
        "Max Value",
    ]
    ws_features.append(feature_headers)
    _style_header(ws_features, len(feature_headers))

    for spec in specs:
        for feat in spec.features:
            ws_features.append(
                [
                    spec.name,
                    feat.name,
                    feat.data_type,
                    feat.unit or "",
                    "Yes" if feat.is_required else "No",
                    feat.min_value,
                    feat.max_value,
                ]
            )
    _auto_width(ws_features)

    wb.save(filepath)
    return len(specs)


def export_purchase_orders(session: Session, filepath: Union[str, Path]) -> int:
    """Export all purchase orders to Excel. Returns number of rows exported."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Purchase Orders"

    headers = [
        "ID",
        "PO Number",
        "Vendor",
        "Product",
        "Status",
        "Quantity",
        "Unit Price",
        "Currency",
        "Total Amount",
        "Shipping",
        "Tax",
        "Grand Total",
        "Order Date",
        "Expected Delivery",
        "Actual Delivery",
        "Invoice Number",
        "Notes",
    ]
    ws.append(headers)
    _style_header(ws, len(headers))

    pos = PurchaseOrderService.get_all(session)
    for po in pos:
        ws.append(
            [
                po.id,
                po.po_number,
                po.vendor.name,
                po.product.name,
                po.status,
                po.quantity,
                po.unit_price,
                po.currency,
                po.total_amount,
                po.shipping_cost,
                po.tax,
                po.grand_total,
                po.order_date.isoformat() if po.order_date else "",
                po.expected_delivery.isoformat() if po.expected_delivery else "",
                po.actual_delivery.isoformat() if po.actual_delivery else "",
                po.invoice_number or "",
                po.notes or "",
            ]
        )

    _auto_width(ws)
    wb.save(filepath)
    return len(pos)


def export_all(session: Session, filepath: Union[str, Path]) -> dict[str, int]:
    """Export all data to a single Excel file with multiple sheets."""
    wb = Workbook()

    # Remove default sheet
    wb.remove(wb.active)

    counts = {}

    # Brands
    ws = wb.create_sheet("Brands")
    ws.append(["ID", "Name", "Product Count"])
    _style_header(ws, 3)
    brands = BrandService.get_all(session)
    for b in brands:
        ws.append([b.id, b.name, len(b.products)])
    _auto_width(ws)
    counts["brands"] = len(brands)

    # Products
    ws = wb.create_sheet("Products")
    ws.append(["ID", "Name", "Brand", "Category", "Specification"])
    _style_header(ws, 5)
    products = ProductService.get_all(session)
    for p in products:
        ws.append(
            [
                p.id,
                p.name,
                p.brand.name if p.brand else "",
                p.category or "",
                p.specification.name if p.specification else "",
            ]
        )
    _auto_width(ws)
    counts["products"] = len(products)

    # Vendors
    ws = wb.create_sheet("Vendors")
    ws.append(
        [
            "ID",
            "Name",
            "Currency",
            "Discount Code",
            "Discount %",
            "Email",
            "Phone",
            "City",
            "Country",
        ]
    )
    _style_header(ws, 9)
    vendors = VendorService.get_all(session)
    for v in vendors:
        ws.append(
            [
                v.id,
                v.name,
                v.currency,
                v.discount_code,
                v.discount,
                v.email,
                v.phone,
                v.city,
                v.country,
            ]
        )
    _auto_width(ws)
    counts["vendors"] = len(vendors)

    # Quotes
    ws = wb.create_sheet("Quotes")
    ws.append(
        ["ID", "Vendor", "Product", "Price", "Currency", "Shipping", "Total", "Status"]
    )
    _style_header(ws, 8)
    quotes = QuoteService.get_all(session)
    for q in quotes:
        ws.append(
            [
                q.id,
                q.vendor.name,
                q.product.name,
                q.value,
                q.currency,
                q.shipping_cost,
                q.total_cost,
                q.status,
            ]
        )
    _auto_width(ws)
    counts["quotes"] = len(quotes)

    # Specifications
    ws = wb.create_sheet("Specifications")
    ws.append(["ID", "Name", "Description", "Features", "Products"])
    _style_header(ws, 5)
    specs = SpecificationService.get_all(session)
    for s in specs:
        ws.append([s.id, s.name, s.description or "", len(s.features), len(s.products)])
    _auto_width(ws)
    counts["specifications"] = len(specs)

    # Purchase Orders
    ws = wb.create_sheet("Purchase Orders")
    ws.append(
        ["ID", "PO#", "Vendor", "Product", "Qty", "Total", "Status", "Order Date"]
    )
    _style_header(ws, 8)
    pos = PurchaseOrderService.get_all(session)
    for po in pos:
        ws.append(
            [
                po.id,
                po.po_number,
                po.vendor.name,
                po.product.name,
                po.quantity,
                po.grand_total,
                po.status,
                po.order_date.isoformat() if po.order_date else "",
            ]
        )
    _auto_width(ws)
    counts["purchase_orders"] = len(pos)

    wb.save(filepath)
    return counts


# =============================================================================
# Template Generation (for import)
# =============================================================================


def generate_vendor_template(filepath: Union[str, Path]) -> None:
    """Generate an Excel template for vendor import."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Vendors"

    headers = [
        "Name *",
        "Currency",
        "Discount Code",
        "Discount %",
        "URL",
        "Contact Person",
        "Email",
        "Phone",
        "Website",
        "Address Line 1",
        "Address Line 2",
        "City",
        "State",
        "Postal Code",
        "Country",
        "Tax ID",
        "Payment Terms",
    ]
    ws.append(headers)
    _style_header(ws, len(headers))

    # Add example row
    ws.append(
        [
            "Example Vendor",
            "USD",
            "SAVE10",
            "10",
            "https://example.com",
            "John Doe",
            "john@example.com",
            "+1-555-1234",
            "https://example.com",
            "123 Main St",
            "Suite 100",
            "New York",
            "NY",
            "10001",
            "USA",
            "12-3456789",
            "Net 30",
        ]
    )

    # Add currency validation
    currencies = ["USD", "EUR", "GBP", "JPY", "CAD", "AUD", "CHF", "CNY"]
    _add_data_validation(ws, 2, currencies)

    _auto_width(ws)
    wb.save(filepath)


def generate_product_template(
    filepath: Union[str, Path], session: Optional[Session] = None
) -> None:
    """Generate an Excel template for product import."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Products"

    headers = ["Name *", "Brand *", "Category"]
    ws.append(headers)
    _style_header(ws, len(headers))

    ws.append(["Example Product", "Example Brand", "Electronics"])

    # Add brand validation if session provided
    if session:
        brands = [b.name for b in BrandService.get_all(session)]
        if brands:
            _add_data_validation(ws, 2, brands)

    _auto_width(ws)
    wb.save(filepath)


def generate_quote_template(
    filepath: Union[str, Path], session: Optional[Session] = None
) -> None:
    """Generate an Excel template for quote import."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Quotes"

    headers = [
        "Vendor *",
        "Product *",
        "Brand (if new)",
        "Price *",
        "Currency",
        "Shipping Cost",
    ]
    ws.append(headers)
    _style_header(ws, len(headers))

    ws.append(["Amazon", "iPhone 15", "Apple", "999.99", "USD", "0"])

    # Add validations if session provided
    if session:
        vendors = [v.name for v in VendorService.get_all(session)]
        if vendors:
            _add_data_validation(ws, 1, vendors)
        products = [p.name for p in ProductService.get_all(session)]
        if products:
            _add_data_validation(ws, 2, products)

    currencies = ["USD", "EUR", "GBP", "JPY", "CAD", "AUD"]
    _add_data_validation(ws, 5, currencies)

    _auto_width(ws)
    wb.save(filepath)


def generate_specification_template(filepath: Union[str, Path]) -> None:
    """Generate an Excel template for specification import."""
    wb = Workbook()

    # Sheet 1: Specifications
    ws = wb.active
    ws.title = "Specifications"
    ws.append(["Name *", "Description"])
    _style_header(ws, 2)
    ws.append(["Camera Spec", "Specifications for cameras"])
    _auto_width(ws)

    # Sheet 2: Features
    ws_feat = wb.create_sheet("Features")
    ws_feat.append(
        [
            "Specification Name *",
            "Feature Name *",
            "Data Type",
            "Unit",
            "Required",
            "Min Value",
            "Max Value",
        ]
    )
    _style_header(ws_feat, 7)
    ws_feat.append(["Camera Spec", "Resolution", "number", "MP", "Yes", "1", "200"])
    ws_feat.append(["Camera Spec", "Brand", "text", "", "No", "", ""])
    ws_feat.append(["Camera Spec", "Has WiFi", "boolean", "", "No", "", ""])

    _add_data_validation(ws_feat, 3, SPEC_DATA_TYPES)
    _add_data_validation(ws_feat, 5, ["Yes", "No"])

    _auto_width(ws_feat)
    wb.save(filepath)


def generate_purchase_order_template(
    filepath: Union[str, Path], session: Optional[Session] = None
) -> None:
    """Generate an Excel template for purchase order import."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Purchase Orders"

    headers = [
        "PO Number *",
        "Vendor *",
        "Product *",
        "Quantity",
        "Unit Price *",
        "Currency",
        "Shipping Cost",
        "Tax",
        "Order Date",
        "Expected Delivery",
        "Invoice Number",
        "Notes",
        "Status",
    ]
    ws.append(headers)
    _style_header(ws, len(headers))

    today = datetime.date.today().isoformat()
    ws.append(
        [
            f"PO-{datetime.date.today().strftime('%Y%m%d')}-001",
            "Amazon",
            "iPhone 15",
            "1",
            "999.99",
            "USD",
            "0",
            "0",
            today,
            "",
            "",
            "",
            "pending",
        ]
    )

    # Add validations
    if session:
        vendors = [v.name for v in VendorService.get_all(session)]
        if vendors:
            _add_data_validation(ws, 2, vendors)
        products = [p.name for p in ProductService.get_all(session)]
        if products:
            _add_data_validation(ws, 3, products)

    _add_data_validation(ws, 13, PO_STATUSES)

    _auto_width(ws)
    wb.save(filepath)


# =============================================================================
# Import Functions
# =============================================================================


def _parse_date(value: Any) -> datetime.date | None:
    """Parse date from various formats."""
    if value is None or value == "":
        return None
    if isinstance(value, datetime.datetime):
        return value.date()
    if isinstance(value, datetime.date):
        return value
    if isinstance(value, str):
        return datetime.date.fromisoformat(value)
    return None


def _parse_float(value: Any, default: float = 0.0) -> float:
    """Parse float from various formats."""
    if value is None or value == "":
        return default
    return float(value)


def _parse_float_optional(value: Any) -> Optional[float]:
    """Parse float from various formats, returning None if empty."""
    if value is None or value == "":
        return None
    return float(value)


def _parse_int(value: Any, default: int = 1) -> int:
    """Parse int from various formats."""
    if value is None or value == "":
        return default
    return int(value)


def _parse_bool(value: Any) -> bool:
    """Parse boolean from various formats."""
    if value is None:
        return False
    if isinstance(value, bool):
        return value
    if isinstance(value, str):
        return value.lower() in ("yes", "true", "1", "y")
    return bool(value)


def import_vendors(
    session: Session, filepath: Union[str, Path]
) -> tuple[int, list[str]]:
    """
    Import vendors from Excel file.

    Returns: (success_count, list of error messages)
    """
    wb = load_workbook(filepath, data_only=True)
    ws = wb.active

    success = 0
    errors = []

    for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not row or not row[0]:  # Skip empty rows
            continue

        try:
            name = str(row[0]).strip()
            if name.startswith("Example"):  # Skip example rows
                continue

            VendorService.create(
                session,
                name=name,
                currency=str(row[1] or "USD").strip(),
                discount_code=str(row[2]).strip() if row[2] else None,
                discount=_parse_float(row[3]),
                url=str(row[4]).strip() if len(row) > 4 and row[4] else None,
                contact_person=str(row[5]).strip() if len(row) > 5 and row[5] else None,
                email=str(row[6]).strip() if len(row) > 6 and row[6] else None,
                phone=str(row[7]).strip() if len(row) > 7 and row[7] else None,
                website=str(row[8]).strip() if len(row) > 8 and row[8] else None,
                address_line1=str(row[9]).strip() if len(row) > 9 and row[9] else None,
                address_line2=str(row[10]).strip()
                if len(row) > 10 and row[10]
                else None,
                city=str(row[11]).strip() if len(row) > 11 and row[11] else None,
                state=str(row[12]).strip() if len(row) > 12 and row[12] else None,
                postal_code=str(row[13]).strip() if len(row) > 13 and row[13] else None,
                country=str(row[14]).strip() if len(row) > 14 and row[14] else None,
                tax_id=str(row[15]).strip() if len(row) > 15 and row[15] else None,
                payment_terms=str(row[16]).strip()
                if len(row) > 16 and row[16]
                else None,
            )
            success += 1
        except DuplicateError:
            errors.append(f"Row {row_num}: Vendor '{row[0]}' already exists")
        except Exception as e:
            errors.append(f"Row {row_num}: {e}")

    return success, errors


def import_products(
    session: Session, filepath: Union[str, Path]
) -> tuple[int, list[str]]:
    """Import products from Excel file."""
    wb = load_workbook(filepath, data_only=True)
    ws = wb.active

    success = 0
    errors = []

    for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not row or not row[0]:
            continue

        try:
            name = str(row[0]).strip()
            brand_name = str(row[1]).strip() if row[1] else None

            if name.startswith("Example"):
                continue

            if not brand_name:
                errors.append(f"Row {row_num}: Brand name is required")
                continue

            product = ProductService.create(
                session,
                name=name,
                brand_name=brand_name,
            )
            # Set category if provided (ProductService.create doesn't support it)
            if len(row) > 2 and row[2]:
                product.category = str(row[2]).strip()
                session.commit()
            success += 1
        except DuplicateError:
            errors.append(f"Row {row_num}: Product '{row[0]}' already exists")
        except Exception as e:
            errors.append(f"Row {row_num}: {e}")

    return success, errors


def import_quotes(
    session: Session, filepath: Union[str, Path]
) -> tuple[int, list[str]]:
    """Import quotes from Excel file."""
    wb = load_workbook(filepath, data_only=True)
    ws = wb.active

    success = 0
    errors = []

    for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not row or not row[0]:
            continue

        try:
            vendor_name = str(row[0]).strip()
            product_name = str(row[1]).strip()
            brand_name = str(row[2]).strip() if row[2] else None
            price = _parse_float(row[3])

            if vendor_name.startswith("Example") or vendor_name == "Amazon":
                continue

            quote = QuoteService.create(
                session,
                vendor_name=vendor_name,
                product_name=product_name,
                price=price,
                brand_name=brand_name,
            )
            # Set shipping cost if provided (QuoteService.create doesn't support it)
            if len(row) > 5 and row[5]:
                quote.shipping_cost = _parse_float(row[5])
                session.commit()
            success += 1
        except NotFoundError as e:
            errors.append(f"Row {row_num}: {e}")
        except Exception as e:
            errors.append(f"Row {row_num}: {e}")

    return success, errors


def import_specifications(
    session: Session, filepath: Union[str, Path]
) -> tuple[int, list[str]]:
    """Import specifications from Excel file (expects two sheets: Specifications, Features)."""
    wb = load_workbook(filepath, data_only=True)

    success = 0
    errors = []

    # Import specifications from first sheet
    ws_specs = wb["Specifications"] if "Specifications" in wb.sheetnames else wb.active

    specs_created = {}
    for row_num, row in enumerate(
        ws_specs.iter_rows(min_row=2, values_only=True), start=2
    ):
        if not row or not row[0]:
            continue

        try:
            name = str(row[0]).strip()
            if name.startswith("Example") or name == "Camera Spec":
                continue

            spec = SpecificationService.create(
                session,
                name=name,
                description=str(row[1]).strip() if len(row) > 1 and row[1] else None,
            )
            specs_created[name] = spec
            success += 1
        except DuplicateError:
            errors.append(f"Specifications row {row_num}: '{row[0]}' already exists")
        except Exception as e:
            errors.append(f"Specifications row {row_num}: {e}")

    # Import features from second sheet if present
    if "Features" in wb.sheetnames:
        ws_features = wb["Features"]
        for row_num, row in enumerate(
            ws_features.iter_rows(min_row=2, values_only=True), start=2
        ):
            if not row or not row[0] or not row[1]:
                continue

            try:
                spec_name = str(row[0]).strip()
                feature_name = str(row[1]).strip()

                if spec_name == "Camera Spec":
                    continue

                SpecificationService.add_feature(
                    session,
                    spec_name=spec_name,
                    feature_name=feature_name,
                    data_type=str(row[2]).strip()
                    if len(row) > 2 and row[2]
                    else "text",
                    unit=str(row[3]).strip() if len(row) > 3 and row[3] else None,
                    is_required=_parse_bool(row[4]) if len(row) > 4 else False,
                    min_value=_parse_float_optional(row[5])
                    if len(row) > 5 and row[5]
                    else None,
                    max_value=_parse_float_optional(row[6])
                    if len(row) > 6 and row[6]
                    else None,
                )
            except NotFoundError as e:
                errors.append(f"Features row {row_num}: {e}")
            except Exception as e:
                errors.append(f"Features row {row_num}: {e}")

    return success, errors


def import_purchase_orders(
    session: Session, filepath: Union[str, Path]
) -> tuple[int, list[str]]:
    """Import purchase orders from Excel file."""
    wb = load_workbook(filepath, data_only=True)
    ws = wb.active

    success = 0
    errors = []

    for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not row or not row[0]:
            continue

        try:
            po_number = str(row[0]).strip()
            vendor_name = str(row[1]).strip()
            product_name = str(row[2]).strip()

            if po_number.startswith("PO-") and vendor_name == "Amazon":
                continue

            PurchaseOrderService.create(
                session,
                po_number=po_number,
                vendor_name=vendor_name,
                product_name=product_name,
                quantity=_parse_int(row[3]) if len(row) > 3 else 1,
                unit_price=_parse_float(row[4]),
                currency=str(row[5]).strip() if len(row) > 5 and row[5] else "USD",
                shipping_cost=_parse_float(row[6]) if len(row) > 6 else None,
                tax=_parse_float(row[7]) if len(row) > 7 else None,
                order_date=_parse_date(row[8]) if len(row) > 8 else None,
                expected_delivery=_parse_date(row[9]) if len(row) > 9 else None,
                invoice_number=str(row[10]).strip()
                if len(row) > 10 and row[10]
                else None,
                notes=str(row[11]).strip() if len(row) > 11 and row[11] else None,
                status=str(row[12]).strip() if len(row) > 12 and row[12] else "pending",
            )
            success += 1
        except DuplicateError:
            errors.append(f"Row {row_num}: PO '{row[0]}' already exists")
        except NotFoundError as e:
            errors.append(f"Row {row_num}: {e}")
        except Exception as e:
            errors.append(f"Row {row_num}: {e}")

    return success, errors


# =============================================================================
# Convenience function for CLI
# =============================================================================

ENTITY_TYPES = [
    "vendors",
    "products",
    "quotes",
    "specifications",
    "purchase_orders",
    "all",
]

EXPORTERS = {
    "vendors": export_vendors,
    "products": export_products,
    "quotes": export_quotes,
    "specifications": export_specifications,
    "purchase_orders": export_purchase_orders,
    "pos": export_purchase_orders,
    "all": export_all,
}

IMPORTERS = {
    "vendors": import_vendors,
    "products": import_products,
    "quotes": import_quotes,
    "specifications": import_specifications,
    "specs": import_specifications,
    "purchase_orders": import_purchase_orders,
    "pos": import_purchase_orders,
}

TEMPLATE_GENERATORS = {
    "vendors": generate_vendor_template,
    "products": generate_product_template,
    "quotes": generate_quote_template,
    "specifications": generate_specification_template,
    "specs": generate_specification_template,
    "purchase_orders": generate_purchase_order_template,
    "pos": generate_purchase_order_template,
}
