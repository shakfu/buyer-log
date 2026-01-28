"""Tests for Excel import/export functionality."""

import tempfile
from pathlib import Path

import pytest
from openpyxl import load_workbook

from buylog.excel import (
    export_brands,
    export_products,
    export_vendors,
    export_quotes,
    export_specifications,
    export_purchase_orders,
    export_all,
    import_vendors,
    import_products,
    import_quotes,
    import_specifications,
    import_purchase_orders,
    generate_vendor_template,
    generate_product_template,
    generate_quote_template,
    generate_specification_template,
    generate_purchase_order_template,
)
from buylog.models import Vendor, Product
from buylog.services import (
    BrandService,
    ProductService,
    VendorService,
    QuoteService,
    SpecificationService,
    PurchaseOrderService,
)


# =============================================================================
# Export Tests
# =============================================================================

def test_export_brands(dbsession):
    """Test exporting brands to Excel."""
    BrandService.create(dbsession, "Apple")
    BrandService.create(dbsession, "Samsung")

    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
        count = export_brands(dbsession, f.name)
        assert count == 2

        wb = load_workbook(f.name)
        ws = wb.active
        assert ws.title == "Brands"
        assert ws.cell(1, 1).value == "ID"
        assert ws.cell(1, 2).value == "Name"


def test_export_products(dbsession):
    """Test exporting products to Excel."""
    ProductService.create(dbsession, "iPhone 15", "Apple")
    ProductService.create(dbsession, "Galaxy S24", "Samsung")

    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
        count = export_products(dbsession, f.name)
        assert count == 2

        wb = load_workbook(f.name)
        ws = wb.active
        assert ws.title == "Products"


def test_export_vendors(dbsession):
    """Test exporting vendors to Excel."""
    VendorService.create(dbsession, "Amazon", currency="USD", email="test@amazon.com")
    VendorService.create(dbsession, "BestBuy", currency="USD", city="Minneapolis")

    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
        count = export_vendors(dbsession, f.name)
        assert count == 2

        wb = load_workbook(f.name)
        ws = wb.active
        assert ws.title == "Vendors"
        # Check that email column exists (column 7 in full export)
        headers = [ws.cell(1, i).value for i in range(1, 19)]
        assert "Email" in headers


def test_export_quotes(dbsession):
    """Test exporting quotes to Excel."""
    VendorService.create(dbsession, "Amazon", currency="USD")
    ProductService.create(dbsession, "iPhone 15", "Apple")
    QuoteService.create(dbsession, "Amazon", "iPhone 15", 999.99)

    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
        count = export_quotes(dbsession, f.name)
        assert count == 1


def test_export_specifications(dbsession):
    """Test exporting specifications to Excel."""
    spec = SpecificationService.create(dbsession, "Camera Spec", "For cameras")
    SpecificationService.add_feature(dbsession, "Camera Spec", "Resolution", "number", "MP")

    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
        count = export_specifications(dbsession, f.name)
        assert count == 1

        wb = load_workbook(f.name)
        assert "Specifications" in wb.sheetnames
        assert "Features" in wb.sheetnames


def test_export_purchase_orders(dbsession):
    """Test exporting purchase orders to Excel."""
    VendorService.create(dbsession, "Amazon", currency="USD")
    ProductService.create(dbsession, "iPhone 15", "Apple")
    PurchaseOrderService.create(
        dbsession,
        po_number="PO-001",
        vendor_name="Amazon",
        product_name="iPhone 15",
        unit_price=999.99,
    )

    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
        count = export_purchase_orders(dbsession, f.name)
        assert count == 1


def test_export_all(dbsession):
    """Test exporting all data to single Excel file."""
    BrandService.create(dbsession, "Apple")
    VendorService.create(dbsession, "Amazon", currency="USD")
    ProductService.create(dbsession, "iPhone 15", "Apple")
    QuoteService.create(dbsession, "Amazon", "iPhone 15", 999.99)

    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
        counts = export_all(dbsession, f.name)

        assert counts["brands"] == 1
        assert counts["products"] == 1
        assert counts["vendors"] == 1
        assert counts["quotes"] == 1

        wb = load_workbook(f.name)
        assert "Brands" in wb.sheetnames
        assert "Products" in wb.sheetnames
        assert "Vendors" in wb.sheetnames
        assert "Quotes" in wb.sheetnames


# =============================================================================
# Template Tests
# =============================================================================

def test_generate_vendor_template():
    """Test generating vendor import template."""
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
        generate_vendor_template(f.name)

        wb = load_workbook(f.name)
        ws = wb.active
        assert ws.title == "Vendors"
        assert "Name *" in ws.cell(1, 1).value


def test_generate_product_template(dbsession):
    """Test generating product import template."""
    BrandService.create(dbsession, "Apple")

    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
        generate_product_template(f.name, dbsession)

        wb = load_workbook(f.name)
        ws = wb.active
        assert ws.title == "Products"


def test_generate_quote_template(dbsession):
    """Test generating quote import template."""
    VendorService.create(dbsession, "Amazon", currency="USD")

    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
        generate_quote_template(f.name, dbsession)

        wb = load_workbook(f.name)
        ws = wb.active
        assert ws.title == "Quotes"


def test_generate_specification_template():
    """Test generating specification import template."""
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
        generate_specification_template(f.name)

        wb = load_workbook(f.name)
        assert "Specifications" in wb.sheetnames
        assert "Features" in wb.sheetnames


def test_generate_purchase_order_template(dbsession):
    """Test generating purchase order import template."""
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
        generate_purchase_order_template(f.name, dbsession)

        wb = load_workbook(f.name)
        ws = wb.active
        assert ws.title == "Purchase Orders"


# =============================================================================
# Import Tests
# =============================================================================

def test_import_vendors(dbsession):
    """Test importing vendors from Excel."""
    # Generate template and modify it
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
        generate_vendor_template(f.name)

        # Load and add real data
        wb = load_workbook(f.name)
        ws = wb.active
        # Row 2 has example data, add row 3 with real data
        ws.append([
            "Test Vendor", "EUR", "DISC10", "10", "https://test.com",
            "John", "john@test.com", "+1234", "https://test.com",
            "123 St", "", "Berlin", "BE", "10001", "Germany",
            "DE123", "Net 30",
        ])
        wb.save(f.name)

        success, errors = import_vendors(dbsession, f.name)
        assert success == 1
        assert len(errors) == 0

        vendor = Vendor.by_name(dbsession, "Test Vendor")
        assert vendor is not None
        assert vendor.currency == "EUR"
        assert vendor.city == "Berlin"


def test_import_products(dbsession):
    """Test importing products from Excel."""
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
        generate_product_template(f.name, dbsession)

        wb = load_workbook(f.name)
        ws = wb.active
        ws.append(["Test Product", "Test Brand", "Electronics"])
        wb.save(f.name)

        success, errors = import_products(dbsession, f.name)
        assert success == 1

        product = Product.by_name(dbsession, "Test Product")
        assert product is not None
        assert product.brand.name == "Test Brand"


def test_import_quotes(dbsession):
    """Test importing quotes from Excel."""
    VendorService.create(dbsession, "Test Vendor", currency="USD")
    ProductService.create(dbsession, "Test Product", "Test Brand")

    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
        generate_quote_template(f.name, dbsession)

        wb = load_workbook(f.name)
        ws = wb.active
        ws.append(["Test Vendor", "Test Product", "", "199.99", "USD", "10"])
        wb.save(f.name)

        success, errors = import_quotes(dbsession, f.name)
        assert success == 1


def test_import_specifications(dbsession):
    """Test importing specifications from Excel."""
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
        generate_specification_template(f.name)

        wb = load_workbook(f.name)
        ws_specs = wb["Specifications"]
        ws_specs.append(["Test Spec", "Test description"])

        ws_features = wb["Features"]
        ws_features.append(["Test Spec", "Weight", "number", "kg", "Yes", "0", "100"])
        wb.save(f.name)

        success, errors = import_specifications(dbsession, f.name)
        assert success == 1

        spec = SpecificationService.get_by_name(dbsession, "Test Spec")
        assert spec is not None
        assert len(spec.features) == 1
        assert spec.features[0].name == "Weight"


def test_import_purchase_orders(dbsession):
    """Test importing purchase orders from Excel."""
    VendorService.create(dbsession, "Test Vendor", currency="USD")
    ProductService.create(dbsession, "Test Product", "Test Brand")

    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
        generate_purchase_order_template(f.name, dbsession)

        wb = load_workbook(f.name)
        ws = wb.active
        ws.append([
            "PO-TEST-001", "Test Vendor", "Test Product",
            "2", "99.99", "USD",
            "10", "5",
            "2025-01-15", "",
            "INV-001", "Test notes", "pending",
        ])
        wb.save(f.name)

        success, errors = import_purchase_orders(dbsession, f.name)
        assert success == 1

        po = PurchaseOrderService.get_by_po_number(dbsession, "PO-TEST-001")
        assert po is not None
        assert po.quantity == 2
        assert po.unit_price == 99.99


def test_import_duplicate_vendor(dbsession):
    """Test that importing duplicate vendor reports error."""
    VendorService.create(dbsession, "Existing Vendor", currency="USD")

    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
        generate_vendor_template(f.name)

        wb = load_workbook(f.name)
        ws = wb.active
        ws.append(["Existing Vendor", "USD"])
        wb.save(f.name)

        success, errors = import_vendors(dbsession, f.name)
        assert success == 0
        assert len(errors) == 1
        assert "already exists" in errors[0]
